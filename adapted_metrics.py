# Creator of this script: Thai, Long-Yang Jan
# Usage: The nnUNet generates summary.json-files which contain the validation and test metrics
#       The summary of the validation metrics of a fold is located in nnUNetTrainer__nnUNetResEncUNetPlans_24G__2d/fold#/validation//summary.json
#       The summary of the validation metrics of CV is located in nnUNetTrainer__nnUNetResEncUNetPlans_24G__2d/crossval_results_folds_0_1_2_3_4/postprocessed/summary.json
#       The summary of the test metrics is located in .../test_best_configuration/labelsPredict_PP/ss####_##_##/summary_test.json
# python adapted_metrics.py

import json #for reading/writing json files
import numpy as np #for mean and std
import os #for working with file paths
from collections import OrderedDict #keeps dictionary keys in the order the were inserted for nicer json output
from datetime import datetime # for the date
import pandas as pd # for the  excel table
from openpyxl.styles import Border, Side #for creating borderlines in excel
from openpyxl.utils import get_column_letter #for cell width adaption in excel
from openpyxl.styles import Font
from scipy.stats import sem #standard error of the mean

###########################
# to be adapted by user
###########################
# json file that should be analyzed: "summary.json" or "summary_test.json"
input_file = "summary_test.json" 
# rounding the digits after the decimal point for dice, iou, accuracy and ratio
round_digits = 7 # saving excel sheet with default: 5 or 7
#rounding to integer for TP/TN/FP/FN
round_integer = 0
###########################

#reading the input json file
def read_json(input_file):
    try: #try to open json file
        with open(input_file, 'r') as f:
            data = json.load(f) #loading the json file
        return data
    except Exception as e: #if not able to open (file not found or invalid json)
        print(f"Error reading input file: {e}")
        return None

# Function to recursively convert all NumPy types to native Python types
def convert_np_types(data):
    if isinstance(data, dict):
        return {key: convert_np_types(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_np_types(item) for item in data]
    elif isinstance(data, np.generic):
        return data.item()  # Convert np.int64, np.float64 to native Python types (int, float)
    else:
        return data

# analysis of metrics
def analyze_data(data):
    metrics_of_interest = ['Dice', 'IoU', 'FN', 'FP', 'TN', 'TP', 'n_pred', 'n_ref']
    raw_metrics_data = {metric: [] for metric in metrics_of_interest}
    accuracy_data = []
    ratio_data = []
    instances = []
    dataset_id = None
    detector_number = None
    #tracking of NaNs for Dice and IoU and 0s for n_ref that occured in the inputfile 
    nan_dice_indices = []
    nan_iou_indices = []
    zero_nref_indices = []

    #loop for TP, TN, FP, FN, Dice, IoU, Accuracy, n_pred and n_ref
    for case in data.get("metric_per_case", []):
        
        #dataset id
        prediction_file = case.get("prediction_file", "")
        if not dataset_id:
            dataset_id = prediction_file.split("nnUNet_results/")[-1].split("/")[0]  # Extract the dataset_id part
        
        #detector number
        if not detector_number:
            detector_number = prediction_file.split("/")[-2]
        
        # instances/samples
        instance_str = prediction_file.split("/")[-1].split("_")[-1].split(".")[0]
        try:
            instance_int = int(instance_str)
            instances.append(instance_int)
        except ValueError:
            pass  # Skip if not a valid integer
        
        #metrics
        metrics = case.get("metrics", {}).get("1", {})
        TP = metrics.get('TP', 0)
        TN = metrics.get('TN', 0)
        FP = metrics.get('FP', 0)
        FN = metrics.get('FN', 0)
        n_pred = metrics.get('n_pred', 0)
        n_ref = metrics.get('n_ref', 0)

        # Store processed metrics which are given in the json file
        for metric in ["TP", "TN", "FP", "FN", "n_pred", "n_ref"]:
            value = np.nan_to_num(metrics.get(metric, 0))
            raw_metrics_data[metric].append(value)
        # Tracking red-flags, zero n_ref
        if n_ref == 0:
            zero_nref_indices.append(len(instances)-1)
        
        # Handle NaNs for Dice and IoU with conditional replacement
        raw_dice = metrics.get("Dice", np.nan)
        raw_iou = metrics.get("IoU", np.nan)
        # Tracking red-flags for Dice
        if np.isnan(raw_dice):
            nan_dice_indices.append(len(instances)-1)
            if TP == 0 and FP == 0 and FN == 0:
                raw_dice = 1.0
            else:
                raw_dice = 0.0
        # Tracking red-flags for IoU
        if np.isnan(raw_iou):
            nan_iou_indices.append(len(instances)-1)
            if TP == 0 and FP == 0 and FN == 0:
                raw_iou = 1.0
            else:
                raw_iou = 0.0
        #store adapted Dice and IoU
        raw_metrics_data["Dice"].append(raw_dice)
        raw_metrics_data["IoU"].append(raw_iou)
    
        # calulate the accuracy
        accuracy = (TP + TN) / (TP + TN + FP + FN)
        accuracy_data.append(accuracy)

    # calculate mean for n_pred for ratio calculation
    mean_n_pred=np.mean(raw_metrics_data["n_pred"])

    for case in data.get("metric_per_case", []):
        
        #metrics
        metrics = case.get("metrics", {}).get("1", {})
        n_pred = metrics.get('n_pred', 0)
        n_ref = metrics.get('n_ref', 0)

        # calculate the ratio
        if n_ref == 0: # Avoid division by 0
            if n_ref == n_pred: # if both are identical set to be 100%
                ratio = 1
            else:
                ratio = max(1-(n_pred/mean_n_pred),0) # else 100%-[n_pred/mean(n_pred)], ratio>=0
        else: # normal formula for all other cases
            ratio = n_pred / n_ref
        ratio_data.append(ratio)

    #ordered dictionary for results
    ordered_result = OrderedDict()
    ordered_result["detector_number"] = detector_number
    ordered_result["instances_dataset"] = len(instances)
    ordered_result["instances"] = instances

    # calculate statistics for each metric
    # for TP, TN, FP, FN -> rounded to int
    for metric in ["TP", "TN", "FP", "FN"]:
        values = raw_metrics_data[metric]
        if values:
            N = len(values)
            ordered_result[metric] = {
                "N": N,
                "values": values,
                "mean": round(np.mean(values),round_integer),
                "std": round(np.std(values, ddof=1),round_integer),
                "SEM": round(sem(values, ddof=1), round_integer),
                "median": round(np.median(values),round_integer),
                "min": round(np.min(values),round_integer),
                "max": round(np.max(values),round_integer)
            }

    if accuracy_data:
        N = len(accuracy_data)
        ordered_result["Accuracy"] = {
            "N": N,
            "values": accuracy_data,
            "mean": round(np.mean(accuracy_data),round_digits),
            "std": round(np.std(accuracy_data, ddof=1),round_digits),
            "SEM": round(sem(accuracy_data, ddof=1), round_digits),
            "median": round(np.median(accuracy_data), round_digits),
            "min": round(np.min(accuracy_data), round_digits),
            "max": round(np.max(accuracy_data), round_digits)
        }

    if raw_metrics_data["Dice"]:
        N = len(raw_metrics_data["Dice"])
        ordered_result["Dice"] = {
            "N": N,
            "values": raw_metrics_data["Dice"],
            "mean": round(np.mean(raw_metrics_data["Dice"]),round_digits),
            "std": round(np.std(raw_metrics_data["Dice"], ddof=1),round_digits),
            "SEM": round(sem(raw_metrics_data["Dice"], ddof=1),round_digits),
            "median": round(np.median(raw_metrics_data["Dice"]),round_digits),
            "min": round(np.min(raw_metrics_data["Dice"]),round_digits),
            "max": round(np.max(raw_metrics_data["Dice"]),round_digits)
        }

    if raw_metrics_data["IoU"]:
        N = len(raw_metrics_data["IoU"])
        ordered_result["IoU"] = {
            "N": N,
            "values": raw_metrics_data["IoU"],
            "mean": round(np.mean(raw_metrics_data["IoU"]),round_digits),
            "std": round(np.std(raw_metrics_data["IoU"], ddof=1),round_digits),
            "SEM": round(sem(raw_metrics_data["IoU"], ddof=1),round_digits),
            "median": round(np.median(raw_metrics_data["IoU"]),round_digits),
            "min": round(np.min(raw_metrics_data["IoU"]),round_digits),
            "max": round(np.max(raw_metrics_data["IoU"]),round_digits)
        }

    for metric in ["n_pred", "n_ref"]:
        values = raw_metrics_data[metric]
        if values:
            N = len(values)
            ordered_result[metric] = {
                "N": N,
                "values": values,
                "mean": round(np.mean(values),round_integer),
                "std": round(np.std(values, ddof=1),round_integer),
                "SEM": round(sem(values, ddof=1),round_integer),
                "median": round(np.median(values),round_integer),
                "min": round(np.min(values),round_integer),
                "max": round(np.max(values),round_integer)
            }

    if ratio_data:
        N=len(ratio_data)
        ordered_result["Ratio"] = {
            "N": N,
            "values": ratio_data,
            "mean": round(np.mean(ratio_data),round_digits),
            "std": round(np.std(ratio_data, ddof=1),round_digits),
            "SEM": round(sem(ratio_data, ddof=1),round_digits),
            "median": round(np.median(ratio_data),round_digits),
            "min": round(np.min(ratio_data),round_digits),
            "max": round(np.max(ratio_data),round_digits)
        }

    # Return extra flags for red highlighting
    red_flags = {
        "Dice": nan_dice_indices,
        "IoU": nan_iou_indices,
        "n_ref": zero_nref_indices
    }
    return ordered_result, dataset_id, detector_number, red_flags


##########################
# creating the excel file
##########################
# Define a thin border
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

#borders for tables
def apply_borders(sheet, dataframe, start_row, start_col):
    """
    Apply borders to the entire range of a dataframe written to the sheet
    """
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, 
                               max_row=start_row + len(dataframe.index), 
                               max_col=start_col + len(dataframe.columns) - 1):
        for cell in row:
            cell.border = border
def apply_borders_summary(sheet, dataframe, start_row, start_col):
    """
    Apply borders to the entire range of a dataframe written to the sheet
    """
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, 
                               max_row=start_row + len(dataframe.index), 
                               max_col=start_col + len(dataframe.columns)):
        for cell in row:
            cell.border = border

def create_excel(input_file, analysis_result, dataset_id, detector_number, red_flags):
    try:
        # Prepare the data for the Excel table
        instances = analysis_result["instances"]
        metrics_data = []

        # Prepare the values for the table of indiviudal samples
        for i, instance in enumerate(instances):
            row = {
                "instances": instance,
                "TP": analysis_result["TP"]["values"][i],
                "TN": analysis_result["TN"]["values"][i],
                "FP": analysis_result["FP"]["values"][i],
                "FN": analysis_result["FN"]["values"][i],
                "Accuracy": round(analysis_result["Accuracy"]["values"][i],round_digits) if "Accuracy" in analysis_result else None,
                "Dice": round(analysis_result["Dice"]["values"][i],round_digits) if "Dice" in analysis_result else None,
                "IoU": round(analysis_result["IoU"]["values"][i],round_digits) if "IoU" in analysis_result else None,
                "n_pred": analysis_result["n_pred"]["values"][i] if "n_pred" in analysis_result else None,
                "n_ref": analysis_result["n_ref"]["values"][i] if "n_ref" in analysis_result else None,
                "Ratio": round(analysis_result["Ratio"]["values"][i],round_digits) if "Ratio" in analysis_result else None
            }
            metrics_data.append(row)

        # Create a DataFrame for metrics
        metrics_df = pd.DataFrame(metrics_data)

        # Prepare the Excel file name
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        excel_filename = f"{base_name}_{detector_number}.xlsx"

        # Create an Excel writer object
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # Write the metadata in the first section
            metadata = {
                "Metrics": "considering whole dataset",
                "date of running the script": datetime.today().strftime('%Y-%m-%d'),
                "dataset_id": dataset_id,  # Add dataset_id to metadata
                "detector_number": detector_number,
                "instances_dataset": len(instances),
                "Integers":f"TP/TN/FP/FN/n_ref/n_pred rounded by {round_integer} decimal digits",
                "Floats":f"Accuracy/Dice/IoU/Ratio rounded by {round_digits} decimal digits",
                "Note on Dice": "Dice=2TP/(2TP+FP+FN); if TP=FP=FN=0 -> Dice=NaN -> replaced by Dice=1 (in red)." ,
                "Note on IoU": "IoU=TP/(TP+FP+FN); if TP=FP=FN=0 -> IoU=NaN -> replaced by IoU=1 (in red)",
                "Note on Ratio": "Ratio=n_pred/n_ref; if n_ref=0 -> Ratio=NaN -> replaced by Ratio=1 for n_ref=n_pred, Ratio=1-(n_pred/mean(n_pred))for n_ref!=n_pred (in red)"                
            }

            # Convert metadata to DataFrame and write to the first rows
            metadata_df = pd.DataFrame(list(metadata.items()), columns=["Parameter", "Description"])
            metadata_df.to_excel(writer, sheet_name=f"{detector_number}", startrow=0, index=False)

            # Create a table for summary statistics (mean, std, median, min, max)
            summary_stats_data = []
            row_labels = ["N", "Mean", "Std", "SEM", "Median", "Min", "Max"]

            for metric in ["TP", "TN", "FP", "FN", "Accuracy", "Dice", "IoU", "n_pred", "n_ref", "Ratio"]:
                if metric in analysis_result:
                    stats = analysis_result[metric]
                    summary_stats_data.append([
                        stats['N'],
                        stats["mean"],
                        stats["std"],
                        stats["SEM"],
                        stats["median"],
                        stats["min"],
                        stats["max"]
                    ])

            # Convert the summary statistics data to DataFrame (metrics as rows and statistics as columns)
            summary_stats_df = pd.DataFrame(summary_stats_data, 
                                            columns=row_labels,
                                            index=["TP", "TN", "FP", "FN", "Accuracy", "corrected Dice", "corrected IoU", "n_pred", "n_ref", "Ratio"])

            # Transpose the DataFrame to switch rows and columns (making row labels appear in the first column)
            summary_stats_df = summary_stats_df.transpose()

            # Write the summary statistics table to the same sheet
            summary_stats_df.to_excel(writer, sheet_name=f"{detector_number}", startrow=len(metadata) + 2, index=True)

            # Write the metrics table to the same sheet
            metrics_df.to_excel(writer, sheet_name=f"{detector_number}", startrow=len(metadata) + len(summary_stats_df) + 4, index=False)

            # Get the active sheet after writing
            sheet = writer.sheets[detector_number]

            # Apply borders to the summary statistics table
            apply_borders_summary(sheet, summary_stats_df, len(metadata) + 3, 1)

            # Apply borders to the metrics table
            apply_borders(sheet, metrics_df, len(metadata) + len(summary_stats_df) + 5, 1)

            # Apply red font color to specific cells based on conditions (NaN for Dice/IoU, 0 for n_ref)
            for col_idx, col_name in enumerate(metrics_df.columns, start=1):  # Excel columns are 1-based
                if col_name in red_flags:
                    for row_offset in range(len(metrics_df)):
                        idx = metrics_df.index[row_offset]
                        if idx in red_flags[col_name]:
                            # Adjust row where red font color should be applied (after metadata and summary statistics)
                            row_num = len(metadata) + len(summary_stats_df) + 5 + row_offset + 1
                            cell = sheet.cell(row=row_num, column=col_idx)
                            cell.font = Font(color="FF0000")  # Set font color to red

            # Apply red font color for Ratio column if n_ref is red
            for row_offset in range(len(metrics_df)):
                idx = metrics_df.index[row_offset]
                if idx in red_flags["n_ref"]:  # Check if n_ref is marked in red for this instance
                    # Apply red color to the corresponding Ratio cell
                    ratio_col_idx = metrics_df.columns.get_loc("Ratio") + 1  # Column index for Ratio (1-based)
                    row_num = len(metadata) + len(summary_stats_df) + 5 + row_offset + 1
                    cell = sheet.cell(row=row_num, column=ratio_col_idx)
                    cell.font = Font(color="FF0000")  # Set font color to red


            #Apply red font color to "TP", "FP" and "FN" if Dice and IoU is red
            # Combine the indices where Dice or IoU are red
            dice_red_rows = set(red_flags.get("Dice", []))
            iou_red_rows = set(red_flags.get("IoU", []))
            rows_to_mark_tp_fp_fn = dice_red_rows.union(iou_red_rows)

            # Get column indices for TP, FP, TN
            tp_col_idx = metrics_df.columns.get_loc("TP") + 1
            fp_col_idx = metrics_df.columns.get_loc("FP") + 1
            fn_col_idx = metrics_df.columns.get_loc("FN") + 1

            for row_offset in range(len(metrics_df)):
                idx = metrics_df.index[row_offset]
                if idx in rows_to_mark_tp_fp_fn:
                    row_num = len(metadata) + len(summary_stats_df) + 5 + row_offset + 1
                    # Color TP cell red
                    cell_tp = sheet.cell(row=row_num, column=tp_col_idx)
                    cell_tp.font = Font(color="FF0000")
                    # Color FP cell red
                    cell_fp = sheet.cell(row=row_num, column=fp_col_idx)
                    cell_fp.font = Font(color="FF0000")
                    # Color FN cell red
                    cell_fn = sheet.cell(row=row_num, column=fn_col_idx)
                    cell_fn.font = Font(color="FF0000")

        print(f"Excel sheet saved to {excel_filename}")
    except Exception as e:
        print(f"Error writing to Excel: {e}")

def main():
    data = read_json(input_file)
    if data is None:
        return
    analysis_result, dataset_id, detector_number, red_flags = analyze_data(data)

    # Creating the Excel file
    create_excel(input_file, analysis_result, dataset_id, detector_number, red_flags)

if __name__ == "__main__":
    main()
